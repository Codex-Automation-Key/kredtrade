from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from django.contrib.auth.models import User
from users.models import Profile, State

class Command(BaseCommand):
    help = 'Adds or updates profiles from an Excel file'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='The path to the Excel file.')

    def handle(self, *args, **kwargs):
        file_path = kwargs['file_path']
        try:
            wb = load_workbook(filename=file_path)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                user, created = User.objects.get_or_create(username=row[0], email=row[1])
                if created:
                    user.set_password('Pass*0502')  # Set a default password or customize as needed
                    user.save()
                
                # Check if the profile exists and update or create accordingly
                profile, created = Profile.objects.update_or_create(
                    user=user,
                    defaults={
                        'company': row[2],
                        'address': row[3],
                        'country': row[4],
                        'est': row[6],
                        'nature': row[7],
                        'status': row[8],
                        'description': row[9],
                        'turnover': row[10],
                        'employee': row[11],
                        'promotor_name': row[12],
                        'promotror_mail': row[13],
                        'promotor_mob': row[14],
                        'auth_name': row[15],
                        'auth_mail': row[16],
                        'auth_mob': row[17],
                    }
                )

            self.stdout.write(self.style.SUCCESS('Successfully added/updated profiles from "%s"' % file_path))
        except Exception as e:
            raise CommandError('Error adding/updating profiles from "%s": %s' % (file_path, e))
