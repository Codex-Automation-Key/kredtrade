from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from users.models import State  # Replace 'yourapp' with your actual app name

class Command(BaseCommand):
    help = 'Import states from an Excel file'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='The path to the Excel file.')

    def handle(self, *args, **options):
        file_path = options['file_path']
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row is the header
            state_name = row[0]  # Adjust the index based on the structure of your Excel
            state, created = State.objects.get_or_create(name=state_name)

            if created:
                self.stdout.write(self.style.SUCCESS(f'Successfully added state: {state_name}'))
            else:
                self.stdout.write(self.style.WARNING(f'State already exists: {state_name}'))

        self.stdout.write(self.style.SUCCESS('Finished importing states'))
